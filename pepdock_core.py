from __future__ import annotations

import itertools
import math
import re
from dataclasses import dataclass, field
from io import StringIO
from pathlib import Path
from typing import Iterable

from rdkit import Chem
from rdkit.Chem import AllChem, Descriptors, rdchem

import openmm
from openmm import app, unit

import openpyxl


MAX_PEPTIDE_LENGTH = 20
RANDOM_SEED = 42
RDKit_MAX_ITERS = 2000
OPENMM_MAX_ITERS = 2000

STANDARD_AMINO_ACIDS = "ACDEFGHIKLMNPQRSTVWY"

AA_CLASSES = {
    "Aromatic": "FWY",
    "Acidic": "DE",
    "Basic": "KRH",
    "Polar neutral": "STNQ",
    "Hydrophobic": "AVLIM",
    "Small": "GAS",
    "Rigid / Proline": "P",
    "Reactive / Cys": "C",
    "All standard": STANDARD_AMINO_ACIDS,
}

FORCE_FIELD_PRESETS = {
    "amber19_gbn2": {
        "label": "Amber19 + GBn2 implicit solvent",
        "files": ("amber19-all.xml", "implicit/gbn2.xml"),
        "system_kwargs": {
            "nonbondedMethod": app.NoCutoff,
            "constraints": None,
            "soluteDielectric": 1.0,
            "solventDielectric": 78.5,
        },
    },
    "charmm36": {
        "label": "CHARMM36 vacuum",
        "files": ("charmm36.xml",),
        "system_kwargs": {
            "nonbondedMethod": app.NoCutoff,
            "constraints": None,
        },
    },
}

PUBLICATION_DESCRIPTOR_HEADERS = (
    "Serial",
    "Sequence",
    "Length",
    "Source",
    "Selected",
    "Status",
    "Rank",
    "Score_0_100",
    "ScoreBand",
    "ScoreNotes",
    "SolubilityScore_0_100",
    "SolubilityClass",
    "SolubilityNotes",
    "GRAVY",
    "IsoelectricPoint_est",
    "ChargedFraction",
    "PolarFraction",
    "HydrophobicPatchMax",
    "Build3DRecommended",
    "MolWt_Da",
    "ExactMass_Da",
    "FormalCharge",
    "EstimatedCharge_pH7_4",
    "TPSA_A2",
    "HBD",
    "HBA",
    "RotatableBonds",
    "HeavyAtoms",
    "cLogP_RDKit",
    "HydrophobicFraction",
    "AromaticCount",
    "AcidicCount",
    "BasicCount",
    "PolarCount",
    "SmallCount",
)

HYDROPHOBIC_RESIDUES = set("AVLIMFWYP")
HYDROPHOBIC_PATCH_RESIDUES = set("AVLIMFWY")
POLAR_RESIDUES = set("STNQCY")
SOLUBILITY_POLAR_RESIDUES = set("DEKRHNQST")
CHARGED_RESIDUES = set("DEKRH")
SMALL_RESIDUES = set("GAS")
REACTIVE_RESIDUES = set("C")

KYTE_DOOLITTLE = {
    "A": 1.8,
    "C": 2.5,
    "D": -3.5,
    "E": -3.5,
    "F": 2.8,
    "G": -0.4,
    "H": -3.2,
    "I": 4.5,
    "K": -3.9,
    "L": 3.8,
    "M": 1.9,
    "N": -3.5,
    "P": -1.6,
    "Q": -3.5,
    "R": -4.5,
    "S": -0.8,
    "T": -0.7,
    "V": 4.2,
    "W": -0.9,
    "Y": -1.3,
}

SIDECHAIN_PKA = {
    "C": 8.3,
    "D": 3.9,
    "E": 4.1,
    "H": 6.0,
    "K": 10.5,
    "R": 12.5,
    "Y": 10.1,
}
N_TERM_PKA = 9.6
C_TERM_PKA = 2.4
DEFAULT_SOLUBILITY_THRESHOLD = 55
MAX_QC_ISSUES = 12


@dataclass(frozen=True)
class PeptideRecord:
    serial: int
    sequence: str
    source: str = "manual"
    values: tuple = field(default_factory=tuple)
    selected: bool = True
    status: str = "pending"
    reason: str = ""


@dataclass(frozen=True)
class BuildSettings:
    output_dir: Path
    formats: tuple[str, ...] = ("pdb",)
    pdb_record_type: str = "ATOM"
    force_field: str = "amber19_gbn2"
    source_name: str = "PepDock Forge"


@dataclass(frozen=True)
class BuiltStructure:
    record: PeptideRecord
    pdb_block: str
    rdkit_mol: Chem.Mol
    energy_kj_mol: float
    variants: tuple[str, ...]
    force_field: str


@dataclass(frozen=True)
class BuildResult:
    record: PeptideRecord
    ok: bool
    files: tuple[Path, ...] = field(default_factory=tuple)
    energy_kj_mol: float | None = None
    message: str = ""


@dataclass(frozen=True)
class PdbAtom:
    serial: int
    name: str
    residue_name: str
    residue_number: int
    element: str
    x: float
    y: float
    z: float


@dataclass(frozen=True)
class PdbStructure:
    atoms: tuple[PdbAtom, ...]
    bonds: tuple[tuple[int, int], ...]


def normalize_sequence(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).upper()


def validate_sequence(sequence: str, max_length: int = MAX_PEPTIDE_LENGTH) -> str:
    sequence = normalize_sequence(sequence)
    if not sequence:
        raise ValueError("Sequence is empty.")
    if len(sequence) > max_length:
        raise ValueError(f"Sequence has {len(sequence)} residues; maximum is {max_length}.")

    invalid = sorted(set(sequence) - set(STANDARD_AMINO_ACIDS))
    if invalid:
        raise ValueError(f"Unsupported amino acid code(s): {', '.join(invalid)}.")
    return sequence


def parse_serial_number(value: object, fallback: int) -> int:
    if isinstance(value, bool):
        return fallback
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip() if value is not None else ""
    if re.fullmatch(r"\d+(\.0+)?", text):
        return int(float(text))
    return fallback


def load_peptide_rows_from_xlsx(xlsx_path: str | Path) -> tuple[tuple[str, ...], list[PeptideRecord]]:
    path = Path(xlsx_path)
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    records: list[PeptideRecord] = []
    header: tuple[str, ...] = ()

    try:
        worksheet = workbook.active
        for raw_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = tuple(row)
            if len(values) < 2:
                continue

            sequence = normalize_sequence(values[1])
            try:
                sequence = validate_sequence(sequence)
            except ValueError:
                if raw_index <= 5 and any(cell not in (None, "") for cell in values):
                    header = tuple(str(cell) if cell is not None else "" for cell in values)
                continue

            records.append(
                PeptideRecord(
                    serial=parse_serial_number(values[0], len(records) + 1),
                    sequence=sequence,
                    source=path.name,
                    values=values,
                )
            )
    finally:
        workbook.close()

    if not header:
        header = ("Serial", "Sequence", "Molecular Weight", "Net Charge", "Solubility", "Isoelectric Point")
    return header, records


def export_library_to_xlsx(
    xlsx_path: str | Path,
    records: Iterable[PeptideRecord],
    imported_header: tuple[str, ...] = (),
) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Peptide Library"

    base_header = list(imported_header) if imported_header else ["Serial", "Sequence"]
    metadata = ["PepDock Selected", "PepDock Status", "PepDock Reason", "PepDock Source"]
    worksheet.append(base_header + metadata)

    for record in records:
        values = list(record.values) if record.values else [record.serial, record.sequence]
        if len(values) < len(base_header):
            values.extend([""] * (len(base_header) - len(values)))
        worksheet.append(values[: len(base_header)] + [record.selected, record.status, record.reason, record.source])

    workbook.save(xlsx_path)


def generate_combinatorial_library(
    position_choices: list[str],
    max_count: int = 10000,
    start_serial: int = 1,
) -> list[PeptideRecord]:
    if not position_choices:
        raise ValueError("No positions were provided.")
    if len(position_choices) > MAX_PEPTIDE_LENGTH:
        raise ValueError(f"Maximum peptide length is {MAX_PEPTIDE_LENGTH}.")

    clean_choices: list[str] = []
    for index, choices in enumerate(position_choices, start=1):
        unique = "".join(dict.fromkeys(normalize_sequence(choices)))
        validate_sequence(unique, max_length=len(unique))
        clean_choices.append(unique)
        if not unique:
            raise ValueError(f"Position {index} has no amino acid choices.")

    total = combination_count(clean_choices)
    if total > max_count:
        raise ValueError(f"Library would contain {total} peptides; current cap is {max_count}.")

    records = []
    for offset, parts in enumerate(itertools.product(*clean_choices)):
        sequence = "".join(parts)
        records.append(
            PeptideRecord(
                serial=start_serial + offset,
                sequence=sequence,
                source="generated",
                values=(start_serial + offset, sequence),
            )
        )
    return records


def combination_count(position_choices: Iterable[str]) -> int:
    total = 1
    for choices in position_choices:
        total *= len(normalize_sequence(choices))
    return total


def force_field_label(mode: str) -> str:
    preset = FORCE_FIELD_PRESETS.get(mode)
    if preset is None:
        return mode
    return str(preset["label"])


def peptide_charge_at_ph(sequence: str, ph: float) -> float:
    sequence = validate_sequence(sequence)
    charge = 1.0 / (1.0 + (10 ** (ph - N_TERM_PKA)))
    charge -= 1.0 / (1.0 + (10 ** (C_TERM_PKA - ph)))

    for aa in sequence:
        pka = SIDECHAIN_PKA.get(aa)
        if pka is None:
            continue
        if aa in {"K", "R", "H"}:
            charge += 1.0 / (1.0 + (10 ** (ph - pka)))
        else:
            charge -= 1.0 / (1.0 + (10 ** (pka - ph)))
    return charge


def estimate_charge_ph74(sequence: str) -> float:
    return round(peptide_charge_at_ph(sequence, 7.4), 2)


def estimate_isoelectric_point(sequence: str) -> float:
    sequence = validate_sequence(sequence)
    low = 0.0
    high = 14.0
    for _ in range(50):
        midpoint = (low + high) / 2.0
        if peptide_charge_at_ph(sequence, midpoint) > 0:
            low = midpoint
        else:
            high = midpoint
    return round((low + high) / 2.0, 2)


def gravy(sequence: str) -> float:
    sequence = validate_sequence(sequence)
    return round(sum(KYTE_DOOLITTLE[aa] for aa in sequence) / len(sequence), 3)


def max_hydrophobic_patch(sequence: str) -> int:
    sequence = validate_sequence(sequence)
    longest = 0
    current = 0
    for aa in sequence:
        if aa in HYDROPHOBIC_PATCH_RESIDUES:
            current += 1
            longest = max(longest, current)
        else:
            current = 0
    return longest


def solubility_class(score: float) -> str:
    if score >= 75:
        return "High"
    if score >= DEFAULT_SOLUBILITY_THRESHOLD:
        return "Medium"
    if score >= 35:
        return "Low"
    return "Poor"


def solubility_profile(sequence: str) -> dict[str, object]:
    sequence = validate_sequence(sequence)
    length = len(sequence)
    net_charge = estimate_charge_ph74(sequence)
    abs_charge = abs(net_charge)
    gravy_value = gravy(sequence)
    p_i = estimate_isoelectric_point(sequence)
    charged_fraction = sum(1 for aa in sequence if aa in CHARGED_RESIDUES) / length
    polar_fraction = sum(1 for aa in sequence if aa in SOLUBILITY_POLAR_RESIDUES) / length
    hydrophobic_patch = max_hydrophobic_patch(sequence)
    aromatic_count = sum(sequence.count(aa) for aa in "FWY")
    cysteine_count = sequence.count("C")

    score = 60.0
    notes: list[str] = []

    if gravy_value <= -1.0:
        score += 18
        notes.append("hydrophilic GRAVY")
    elif gravy_value <= 0.0:
        score += 12
        notes.append("favorable GRAVY")
    elif gravy_value <= 0.8:
        score += 4
        notes.append("moderate GRAVY")
    elif gravy_value <= 1.5:
        score -= 10
        notes.append("hydrophobic GRAVY")
    else:
        score -= 22
        notes.append("very hydrophobic GRAVY")

    if abs_charge >= 3.0:
        score += 18
        notes.append("strong net charge")
    elif abs_charge >= 1.0:
        score += 10
        notes.append("useful net charge")
    elif abs_charge >= 0.5:
        score += 5
        notes.append("weak net charge")
    else:
        score -= 8
        notes.append("near-neutral net charge")

    if charged_fraction >= 0.25:
        score += 10
        notes.append("many charged residues")
    elif charged_fraction >= 0.10:
        score += 5
        notes.append("some charged residues")
    else:
        score -= 8
        notes.append("few charged residues")

    if polar_fraction >= 0.35:
        score += 10
        notes.append("many polar/charged residues")
    elif polar_fraction >= 0.20:
        score += 5
        notes.append("some polar/charged residues")
    elif polar_fraction < 0.10:
        score -= 8
        notes.append("few polar/charged residues")

    if hydrophobic_patch <= 2:
        score += 8
        notes.append("short hydrophobic patches")
    elif hydrophobic_patch == 3:
        score += 2
        notes.append("moderate hydrophobic patch")
    elif hydrophobic_patch == 4:
        score -= 8
        notes.append("long hydrophobic patch")
    else:
        score -= 18
        notes.append("aggregation-prone hydrophobic patch")

    if abs(p_i - 7.4) < 0.75:
        score -= 10
        notes.append("pI near physiological pH")
    elif abs(p_i - 7.4) < 1.5:
        score -= 4
        notes.append("pI close to physiological pH")
    else:
        score += 4
        notes.append("pI separated from physiological pH")

    if length > 18:
        score -= 10
        notes.append("long peptide")
    elif length > 15:
        score -= 5
        notes.append("moderately long peptide")

    if aromatic_count >= 4:
        score -= 8
        notes.append("high aromatic load")
    elif aromatic_count >= 2:
        score -= 3
        notes.append("moderate aromatic load")

    if cysteine_count:
        score -= 4 * cysteine_count
        notes.append("cysteine oxidation risk")

    score = max(0.0, min(100.0, round(score, 2)))
    return {
        "score": score,
        "class": solubility_class(score),
        "notes": tuple(dict.fromkeys(notes)),
        "gravy": gravy_value,
        "pI": p_i,
        "charge": net_charge,
        "charged_fraction": round(charged_fraction, 3),
        "polar_fraction": round(polar_fraction, 3),
        "hydrophobic_patch": hydrophobic_patch,
    }


def solubility_build_recommended(
    record: PeptideRecord,
    min_score: int = DEFAULT_SOLUBILITY_THRESHOLD,
) -> bool:
    if record.status == "excluded":
        return False
    if not record.selected:
        return False
    return float(solubility_profile(record.sequence)["score"]) >= min_score


def apply_solubility_gate(
    records: Iterable[PeptideRecord],
    min_score: int = DEFAULT_SOLUBILITY_THRESHOLD,
) -> list[PeptideRecord]:
    gated: list[PeptideRecord] = []
    threshold = max(0, min(100, int(min_score)))
    for record in records:
        if record.status == "excluded":
            gated.append(record)
            continue

        profile = solubility_profile(record.sequence)
        score = float(profile["score"])
        if score >= threshold:
            status = "pending" if record.status == "low_solubility" else record.status
            reason = "" if record.status == "low_solubility" else record.reason
            gated.append(
                PeptideRecord(
                    serial=record.serial,
                    sequence=record.sequence,
                    source=record.source,
                    values=record.values,
                    selected=True,
                    status=status,
                    reason=reason,
                )
            )
            continue

        notes = "; ".join(profile["notes"])
        reason = f"Solubility {score:g} ({profile['class']}) below {threshold}: {notes}"
        status = "low_solubility" if record.status in {"pending", "low_solubility"} else record.status
        gated.append(
            PeptideRecord(
                serial=record.serial,
                sequence=record.sequence,
                source=record.source,
                values=record.values,
                selected=False,
                status=status,
                reason=reason,
            )
        )
    return gated


def score_band(score: float) -> str:
    if score >= 80:
        return "High"
    if score >= 60:
        return "Medium"
    if score >= 40:
        return "Low"
    return "Review"


def peptide_rank_score(record: PeptideRecord) -> tuple[float, str, tuple[str, ...]]:
    sequence = validate_sequence(record.sequence)
    reasons: list[str] = []
    warnings: list[str] = []

    if record.status == "excluded" or not record.selected:
        warnings.append("Excluded or not selected")

    length = len(sequence)
    charge = estimate_charge_ph74(sequence)
    hydrophobic_fraction = sum(1 for aa in sequence if aa in HYDROPHOBIC_RESIDUES) / length
    aromatic_count = sum(sequence.count(aa) for aa in "FWY")
    polar_count = sum(1 for aa in sequence if aa in POLAR_RESIDUES)
    small_count = sum(1 for aa in sequence if aa in SMALL_RESIDUES)
    reactive_count = sum(1 for aa in sequence if aa in REACTIVE_RESIDUES)
    proline_fraction = sequence.count("P") / length
    solubility = solubility_profile(sequence)

    score = 35.0

    if 5 <= length <= 12:
        score += 14
        reasons.append("preferred peptide length")
    elif 13 <= length <= MAX_PEPTIDE_LENGTH:
        score += 8
        reasons.append("long but supported peptide")
    else:
        score += 4
        warnings.append("very short peptide")

    if -2.0 <= charge <= 3.0:
        score += 14
        reasons.append("balanced estimated charge")
    elif -4.0 <= charge <= 5.0:
        score += 8
        warnings.append("moderate charge")
    else:
        score += 2
        warnings.append("extreme estimated charge")

    if 0.25 <= hydrophobic_fraction <= 0.65:
        score += 14
        reasons.append("balanced hydrophobic fraction")
    elif 0.15 <= hydrophobic_fraction <= 0.75:
        score += 7
        warnings.append("borderline hydrophobic balance")
    else:
        score += 1
        warnings.append("extreme hydrophobic balance")

    if aromatic_count:
        score += min(aromatic_count, 3) * 3
        reasons.append("aromatic contact potential")

    if polar_count:
        score += min(polar_count, 4) * 2
        reasons.append("polar H-bond potential")

    if small_count:
        score += min(small_count, 3)

    solubility_score = float(solubility["score"])
    if solubility_score >= 75:
        score += 10
        reasons.append("high solubility score")
    elif solubility_score >= DEFAULT_SOLUBILITY_THRESHOLD:
        score += 5
        reasons.append("acceptable solubility score")
    elif solubility_score < 35:
        score -= 14
        warnings.append("poor solubility score")
    else:
        score -= 8
        warnings.append("low solubility score")

    if reactive_count:
        score -= reactive_count * 6
        warnings.append("reactive cysteine")

    if proline_fraction > 0.25:
        score -= 6
        warnings.append("high proline fraction")

    if record.status == "excluded" or not record.selected:
        score = min(score, 35)

    score = max(0.0, min(100.0, round(score, 2)))
    notes = tuple(dict.fromkeys(reasons + warnings))
    return score, score_band(score), notes


def descriptor_row(record: PeptideRecord) -> dict[str, object]:
    sequence = validate_sequence(record.sequence)
    mol = Chem.MolFromSequence(sequence)
    if mol is None:
        raise ValueError(f"RDKit could not build peptide sequence: {sequence}")

    length = len(sequence)
    formal_charge = sum(atom.GetFormalCharge() for atom in mol.GetAtoms())
    hydrophobic_fraction = sum(1 for aa in sequence if aa in HYDROPHOBIC_RESIDUES) / length
    score, band, notes = peptide_rank_score(record)
    solubility = solubility_profile(sequence)

    return {
        "Serial": record.serial,
        "Sequence": sequence,
        "Length": length,
        "Source": record.source,
        "Selected": record.selected,
        "Status": record.status,
        "Rank": "",
        "Score_0_100": score,
        "ScoreBand": band,
        "ScoreNotes": "; ".join(notes),
        "SolubilityScore_0_100": solubility["score"],
        "SolubilityClass": solubility["class"],
        "SolubilityNotes": "; ".join(solubility["notes"]),
        "GRAVY": solubility["gravy"],
        "IsoelectricPoint_est": solubility["pI"],
        "ChargedFraction": solubility["charged_fraction"],
        "PolarFraction": solubility["polar_fraction"],
        "HydrophobicPatchMax": solubility["hydrophobic_patch"],
        "Build3DRecommended": solubility_build_recommended(record),
        "MolWt_Da": round(Descriptors.MolWt(mol), 3),
        "ExactMass_Da": round(Descriptors.ExactMolWt(mol), 3),
        "FormalCharge": formal_charge,
        "EstimatedCharge_pH7_4": estimate_charge_ph74(sequence),
        "TPSA_A2": round(Descriptors.TPSA(mol), 2),
        "HBD": int(Descriptors.NumHDonors(mol)),
        "HBA": int(Descriptors.NumHAcceptors(mol)),
        "RotatableBonds": int(Descriptors.NumRotatableBonds(mol)),
        "HeavyAtoms": int(Descriptors.HeavyAtomCount(mol)),
        "cLogP_RDKit": round(Descriptors.MolLogP(mol), 3),
        "HydrophobicFraction": round(hydrophobic_fraction, 3),
        "AromaticCount": sum(sequence.count(aa) for aa in "FWY"),
        "AcidicCount": sum(sequence.count(aa) for aa in "DE"),
        "BasicCount": sum(sequence.count(aa) for aa in "KRH"),
        "PolarCount": sum(1 for aa in sequence if aa in POLAR_RESIDUES),
        "SmallCount": sum(1 for aa in sequence if aa in SMALL_RESIDUES),
    }


def descriptor_rows(records: Iterable[PeptideRecord], ranked: bool = True) -> list[dict[str, object]]:
    rows = [descriptor_row(record) for record in records]
    if not ranked:
        return rows

    rows = sorted(
        rows,
        key=lambda row: (
            -float(row["Score_0_100"]),
            -float(row["SolubilityScore_0_100"]),
            int(row["Length"]),
            str(row["Sequence"]),
            int(row["Serial"]),
        ),
    )
    for rank, row in enumerate(rows, start=1):
        row["Rank"] = rank
    return rows


def export_descriptor_report_xlsx(
    xlsx_path: str | Path,
    records: Iterable[PeptideRecord],
    ranked: bool = True,
) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Publication Descriptors"
    worksheet.append(list(PUBLICATION_DESCRIPTOR_HEADERS))
    for row in descriptor_rows(records, ranked=ranked):
        worksheet.append([row.get(header, "") for header in PUBLICATION_DESCRIPTOR_HEADERS])
    methods = workbook.create_sheet("Methods")
    methods.append(["Method Step"])
    for line in report_methods_text():
        methods.append([line])
    workbook.save(xlsx_path)


def report_methods_text() -> list[str]:
    force_fields = ", ".join(force_field_label(key) for key in FORCE_FIELD_PRESETS)
    return [
        "Peptide sequences were validated against the 20 standard amino-acid one-letter codes.",
        f"The maximum supported peptide length was {MAX_PEPTIDE_LENGTH} residues.",
        "Initial peptide structures were generated with RDKit from sequence, hydrogens were added, and conformers were embedded with ETKDG using a fixed random seed.",
        "RDKit peptide templates were normalized to standard peptide atom naming, including isoleucine side-chain atom labels.",
        "RDKit export templates used pH 7.4-like protonation for N-termini, C-termini, Asp/Glu side chains, Lys side chains, and Arg side chains.",
        "RDKit seed conformers were pre-minimized with MMFF94 when available, otherwise UFF was used.",
        f"OpenMM minimization was available through these presets: {force_fields}.",
        "Amber19 used amber19-all.xml with implicit/gbn2.xml, NoCutoff nonbonded treatment, solute dielectric 1.0, and solvent dielectric 78.5.",
        "CHARMM36 used charmm36.xml with NoCutoff nonbonded treatment and no implicit solvent term in this build.",
        "Histidine residues were assigned HID during hydrogen addition for consistency in the current workflow.",
        "MOL2, SDF, and PDBQT exports rebuild explicit hydrogens from the minimized heavy-atom geometry and pass a bonded-distance quality check before files are written.",
        "The ranking score is a transparent heuristic from 0 to 100, not a docking affinity or binding free energy.",
        "Ranking rewards supported peptide length, balanced estimated charge at pH 7.4, balanced hydrophobic fraction, aromatic contact potential, and polar H-bond potential.",
        "Ranking penalizes unselected/excluded rows, reactive cysteine, very short peptides, extreme charge, extreme hydrophobicity, low solubility score, and high proline fraction.",
        "Solubility scoring is a sequence-only heuristic from 0 to 100, not an experimental solubility measurement.",
        "Solubility scoring uses Kyte-Doolittle GRAVY, Henderson-Hasselbalch estimated net charge and pI, charged and polar residue fractions, maximum hydrophobic patch length, aromatic load, cysteine count, and peptide length.",
        f"The default 3D-build recommendation threshold is a solubility score of {DEFAULT_SOLUBILITY_THRESHOLD}.",
    ]


def export_publication_report_markdown(
    markdown_path: str | Path,
    records: Iterable[PeptideRecord],
    title: str = "PepDock Forge Peptide Report",
    ranked: bool = True,
) -> None:
    rows = descriptor_rows(records, ranked=ranked)
    selected_count = sum(1 for row in rows if row["Selected"])
    built_count = sum(1 for row in rows if row["Status"] == "built")
    high_count = sum(1 for row in rows if row["ScoreBand"] == "High")
    medium_count = sum(1 for row in rows if row["ScoreBand"] == "Medium")
    high_solubility_count = sum(1 for row in rows if row["SolubilityClass"] == "High")
    medium_solubility_count = sum(1 for row in rows if row["SolubilityClass"] == "Medium")
    build_recommended_count = sum(1 for row in rows if row["Build3DRecommended"])
    lines = [
        f"# {title}",
        "",
        "## Summary",
        "",
        f"- Total peptides: {len(rows)}",
        f"- Selected peptides: {selected_count}",
        f"- Built peptides: {built_count}",
        f"- High-score peptides: {high_count}",
        f"- Medium-score peptides: {medium_count}",
        f"- High-solubility peptides: {high_solubility_count}",
        f"- Medium-solubility peptides: {medium_solubility_count}",
        f"- Recommended for 3D build: {build_recommended_count}",
        f"- Maximum supported peptide length: {MAX_PEPTIDE_LENGTH} amino acids",
        f"- Ranking: {'score-descending' if ranked else 'input order'}",
        "",
        "## Methods",
        "",
        *[f"- {line}" for line in report_methods_text()],
        "",
        "## Descriptor Table",
        "",
        "| " + " | ".join(PUBLICATION_DESCRIPTOR_HEADERS) + " |",
        "| " + " | ".join("---" for _ in PUBLICATION_DESCRIPTOR_HEADERS) + " |",
    ]
    for row in rows:
        values = [str(row.get(header, "")) for header in PUBLICATION_DESCRIPTOR_HEADERS]
        lines.append("| " + " | ".join(value.replace("|", "\\|") for value in values) + " |")
    lines.append("")
    Path(markdown_path).write_text("\n".join(lines), encoding="utf-8", newline="\n")


def parse_pdb_structure(pdb_text: str) -> PdbStructure:
    atoms: list[PdbAtom] = []
    serials = set()
    bonds = set()

    for line in pdb_text.splitlines():
        record = line[:6].strip()
        if record in {"ATOM", "HETATM"}:
            try:
                serial = int(line[6:11])
                residue_number_text = line[22:26].strip()
                atom = PdbAtom(
                    serial=serial,
                    name=line[12:16].strip(),
                    residue_name=line[17:20].strip(),
                    residue_number=int(residue_number_text) if residue_number_text else 0,
                    element=(line[76:78].strip() or line[12:16].strip()[0]).upper(),
                    x=float(line[30:38]),
                    y=float(line[38:46]),
                    z=float(line[46:54]),
                )
            except (IndexError, ValueError):
                continue
            atoms.append(atom)
            serials.add(serial)
        elif record == "CONECT":
            try:
                source = int(line[6:11])
            except ValueError:
                continue
            for start in range(11, len(line), 5):
                target_text = line[start : start + 5].strip()
                if not target_text:
                    continue
                try:
                    target = int(target_text)
                except ValueError:
                    continue
                if source != target:
                    bonds.add(tuple(sorted((source, target))))

    clean_bonds = tuple(sorted(pair for pair in bonds if pair[0] in serials and pair[1] in serials))
    return PdbStructure(atoms=tuple(atoms), bonds=clean_bonds)


def load_pdb_structure(path: str | Path) -> PdbStructure:
    return parse_pdb_structure(Path(path).read_text(encoding="utf-8", errors="replace"))


def build_seed_molecule(sequence: str) -> Chem.Mol:
    sequence = validate_sequence(sequence)
    template = Chem.MolFromSequence(sequence)
    if template is None:
        raise ValueError(f"RDKit could not build peptide sequence: {sequence}")

    apply_peptide_protonation_ph74(template)
    mol = Chem.AddHs(template)
    standardize_peptide_atom_names(mol)
    params = AllChem.ETKDGv3()
    params.randomSeed = RANDOM_SEED
    result = AllChem.EmbedMolecule(mol, params)
    if result != 0:
        fallback = AllChem.ETKDG()
        fallback.randomSeed = RANDOM_SEED
        result = AllChem.EmbedMolecule(mol, fallback)
    if result != 0:
        result = AllChem.EmbedMolecule(
            mol,
            randomSeed=RANDOM_SEED,
            useRandomCoords=True,
            maxAttempts=1000,
            boxSizeMult=10.0,
        )
    if result != 0:
        raise RuntimeError("RDKit conformer generation failed.")

    if AllChem.MMFFHasAllMoleculeParams(mol):
        AllChem.MMFFOptimizeMolecule(mol, mmffVariant="MMFF94", maxIters=RDKit_MAX_ITERS)
    elif AllChem.UFFHasAllMoleculeParams(mol):
        AllChem.UFFOptimizeMolecule(mol, maxIters=RDKit_MAX_ITERS)
    else:
        raise RuntimeError("Neither MMFF94 nor UFF can parameterize the RDKit seed molecule.")
    return mol


def apply_peptide_protonation_ph74(mol: Chem.Mol) -> None:
    residues: dict[tuple[str, int], dict[str, Chem.Atom]] = {}
    for atom in mol.GetAtoms():
        info = atom.GetPDBResidueInfo()
        if info is None:
            continue
        key = (info.GetResidueName().strip(), info.GetResidueNumber())
        residues.setdefault(key, {})[info.GetName().strip()] = atom

    if not residues:
        return

    first_residue_number = min(number for _name, number in residues)
    for (_residue_name, residue_number), atoms in residues.items():
        if residue_number == first_residue_number and "N" in atoms:
            atoms["N"].SetFormalCharge(1)
        if "OXT" in atoms:
            atoms["OXT"].SetFormalCharge(-1)

        if "OD2" in atoms and "OD1" in atoms:
            atoms["OD2"].SetFormalCharge(-1)
        if "OE2" in atoms and "OE1" in atoms:
            atoms["OE2"].SetFormalCharge(-1)
        if "NZ" in atoms:
            atoms["NZ"].SetFormalCharge(1)
        if "NH2" in atoms and "NH1" in atoms:
            atoms["NH2"].SetFormalCharge(1)

    mol.UpdatePropertyCache(strict=False)


def standardize_peptide_atom_names(mol: Chem.Mol) -> None:
    """Normalize RDKit peptide atom labels to the residue topology OpenMM expects."""
    residues: dict[tuple[str, int], dict[str, Chem.Atom]] = {}
    for atom in mol.GetAtoms():
        info = atom.GetPDBResidueInfo()
        if info is None:
            continue
        key = (info.GetResidueName().strip(), info.GetResidueNumber())
        residues.setdefault(key, {})[info.GetName().strip()] = atom

    for (residue_name, _residue_number), atoms in residues.items():
        if residue_name != "ILE":
            continue
        cg1 = atoms.get("CG1")
        cg2 = atoms.get("CG2")
        if cg1 is None or cg2 is None:
            continue
        cg1_neighbors = {neighbor.GetPDBResidueInfo().GetName().strip() for neighbor in cg1.GetNeighbors() if neighbor.GetPDBResidueInfo()}
        if "CD1" not in cg1_neighbors:
            _set_pdb_atom_name(cg1, "CG2")
            _set_pdb_atom_name(cg2, "CG1")


def _set_pdb_atom_name(atom: Chem.Atom, name: str) -> None:
    info = atom.GetPDBResidueInfo()
    if info is not None:
        info.SetName(f"{name:<4}"[:4])


def _format_atom_name(name: str, element: str) -> str:
    name = name[:4]
    if len(element.strip()) == 1 and len(name.strip()) < 4:
        return f" {name.strip():<3}"[:4]
    return f"{name:<4}"[:4]


def _pdb_atom_line(
    serial: int,
    record_type: str,
    atom_name: str,
    residue_name: str,
    chain_id: str,
    residue_number: int,
    x: float,
    y: float,
    z: float,
    element: str,
) -> str:
    return (
        f"{record_type:<6}{serial:5d} {_format_atom_name(atom_name, element)} {residue_name:>3} "
        f"{chain_id[:1] or 'A'}{residue_number:4d}    "
        f"{x:8.3f}{y:8.3f}{z:8.3f}{1.00:6.2f}{0.00:6.2f}          {element:>2}  "
    )


def rdkit_heavy_pdb_block(mol: Chem.Mol, record_type: str = "ATOM", include_conect: bool = True) -> str:
    conformer = mol.GetConformer()
    lines = []
    serial_by_index = {}
    serial = 1

    for atom in mol.GetAtoms():
        if atom.GetSymbol() == "H":
            continue
        info = atom.GetPDBResidueInfo()
        if info is None:
            raise ValueError("RDKit peptide atom is missing PDB residue metadata.")
        pos = conformer.GetAtomPosition(atom.GetIdx())
        serial_by_index[atom.GetIdx()] = serial
        lines.append(
            _pdb_atom_line(
                serial=serial,
                record_type=record_type,
                atom_name=info.GetName().strip(),
                residue_name=info.GetResidueName().strip(),
                chain_id=info.GetChainId() or "A",
                residue_number=info.GetResidueNumber(),
                x=pos.x,
                y=pos.y,
                z=pos.z,
                element=atom.GetSymbol().upper(),
            )
        )
        serial += 1

    if include_conect:
        lines.extend(_rdkit_conect_lines(mol, serial_by_index))
    lines.append("END")
    return "\n".join(lines) + "\n"


def _rdkit_conect_lines(mol: Chem.Mol, serial_by_index: dict[int, int]) -> list[str]:
    adjacency = {serial: [] for serial in serial_by_index.values()}
    for bond in mol.GetBonds():
        begin = bond.GetBeginAtomIdx()
        end = bond.GetEndAtomIdx()
        if begin not in serial_by_index or end not in serial_by_index:
            continue
        begin_serial = serial_by_index[begin]
        end_serial = serial_by_index[end]
        adjacency[begin_serial].append(end_serial)
        adjacency[end_serial].append(begin_serial)

    lines = []
    for serial in sorted(adjacency):
        neighbors = sorted(adjacency[serial])
        for start in range(0, len(neighbors), 4):
            chunk = neighbors[start : start + 4]
            lines.append("CONECT" + f"{serial:5d}" + "".join(f"{neighbor:5d}" for neighbor in chunk))
    return lines


def _openmm_platform() -> openmm.Platform:
    try:
        return openmm.Platform.getPlatformByName("CPU")
    except Exception:
        return openmm.Platform.getPlatformByName("Reference")


def _forcefield_preset(mode: str) -> dict[str, object]:
    try:
        return FORCE_FIELD_PRESETS[mode]
    except KeyError as exc:
        available = ", ".join(force_field_label(key) for key in FORCE_FIELD_PRESETS)
        raise ValueError(f"Unsupported force field '{mode}'. Available options: {available}.") from exc


def _forcefield_for_mode(mode: str) -> app.ForceField:
    preset = _forcefield_preset(mode)
    return app.ForceField(*preset["files"])


def _system_kwargs_for_mode(mode: str) -> dict[str, object]:
    preset = _forcefield_preset(mode)
    return dict(preset["system_kwargs"])


def minimize_with_openmm(seed_mol: Chem.Mol, sequence: str, settings: BuildSettings) -> BuiltStructure:
    heavy_pdb = rdkit_heavy_pdb_block(seed_mol, record_type="ATOM", include_conect=False)
    pdb = app.PDBFile(StringIO(heavy_pdb))
    forcefield = _forcefield_for_mode(settings.force_field)
    modeller = app.Modeller(pdb.topology, pdb.positions)
    variants = ["HID" if residue.name == "HIS" else None for residue in modeller.topology.residues()]
    selected_variants = modeller.addHydrogens(forcefield, pH=7.4, variants=variants)

    system = forcefield.createSystem(modeller.topology, **_system_kwargs_for_mode(settings.force_field))
    integrator = openmm.VerletIntegrator(0.001 * unit.picoseconds)
    simulation = app.Simulation(modeller.topology, system, integrator, _openmm_platform())
    simulation.context.setPositions(modeller.positions)
    simulation.minimizeEnergy(maxIterations=OPENMM_MAX_ITERS)
    state = simulation.context.getState(positions=True, energy=True)
    positions = state.getPositions()
    energy = state.getPotentialEnergy().value_in_unit(unit.kilojoule_per_mole)
    pdb_block = openmm_pdb_block(
        simulation.topology,
        positions,
        sequence=sequence,
        source_name=settings.source_name,
        force_field=force_field_label(settings.force_field),
        variants=tuple(variant or "standard" for variant in selected_variants),
        energy_kj_mol=energy,
        record_type=settings.pdb_record_type,
    )

    rdkit_export_mol = update_rdkit_heavy_positions(seed_mol, simulation.topology, positions)
    return BuiltStructure(
        record=PeptideRecord(serial=0, sequence=sequence),
        pdb_block=pdb_block,
        rdkit_mol=rdkit_export_mol,
        energy_kj_mol=energy,
        variants=tuple(variant or "standard" for variant in selected_variants),
        force_field=force_field_label(settings.force_field),
    )


def openmm_pdb_block(
    topology: app.Topology,
    positions,
    sequence: str,
    source_name: str,
    force_field: str,
    variants: tuple[str, ...],
    energy_kj_mol: float,
    record_type: str = "ATOM",
) -> str:
    positions_angstrom = positions.value_in_unit(unit.angstrom)
    lines = [
        f"HEADER    PEPDOCK FORGE SEQUENCE {sequence} SOURCE {source_name}",
        "REMARK    Generated by PepDock Forge.",
        "REMARK    Initial conformer: RDKit ETKDGv3, pre-minimized with MMFF94/UFF.",
        f"REMARK    Final minimization: OpenMM {force_field}, energy {energy_kj_mol:.3f} kJ/mol.",
        f"REMARK    Hydrogen variants: {', '.join(variants)}",
    ]

    serial_by_atom = {}
    for serial, (atom, pos) in enumerate(zip(topology.atoms(), positions_angstrom), start=1):
        residue = atom.residue
        residue_number = int(residue.id) if str(residue.id).isdigit() else residue.index + 1
        element = atom.element.symbol.upper() if atom.element is not None else atom.name[0].upper()
        serial_by_atom[atom] = serial
        lines.append(
            _pdb_atom_line(
                serial=serial,
                record_type=record_type,
                atom_name=atom.name,
                residue_name=residue.name,
                chain_id=residue.chain.id or "A",
                residue_number=residue_number,
                x=pos.x,
                y=pos.y,
                z=pos.z,
                element=element,
            )
        )

    lines.extend(_openmm_conect_lines(topology, serial_by_atom))
    lines.append("END")
    return "\n".join(lines) + "\n"


def _openmm_conect_lines(topology: app.Topology, serial_by_atom: dict[object, int]) -> list[str]:
    adjacency = {serial: [] for serial in serial_by_atom.values()}
    for atom_a, atom_b in topology.bonds():
        serial_a = serial_by_atom[atom_a]
        serial_b = serial_by_atom[atom_b]
        adjacency[serial_a].append(serial_b)
        adjacency[serial_b].append(serial_a)

    lines = []
    for serial in sorted(adjacency):
        neighbors = sorted(adjacency[serial])
        for start in range(0, len(neighbors), 4):
            chunk = neighbors[start : start + 4]
            lines.append("CONECT" + f"{serial:5d}" + "".join(f"{neighbor:5d}" for neighbor in chunk))
    return lines


def update_rdkit_heavy_positions(seed_mol: Chem.Mol, topology: app.Topology, positions) -> Chem.Mol:
    mol = Chem.Mol(seed_mol)
    conformer = mol.GetConformer()
    position_map = {}
    positions_angstrom = positions.value_in_unit(unit.angstrom)

    for atom, pos in zip(topology.atoms(), positions_angstrom):
        if atom.element is None or atom.element.symbol.upper() == "H":
            continue
        residue_number = int(atom.residue.id) if str(atom.residue.id).isdigit() else atom.residue.index + 1
        position_map[(residue_number, atom.name.strip())] = pos

    for atom in mol.GetAtoms():
        if atom.GetSymbol() == "H":
            continue
        info = atom.GetPDBResidueInfo()
        if info is None:
            continue
        key = (info.GetResidueNumber(), info.GetName().strip())
        if key in position_map:
            pos = position_map[key]
            conformer.SetAtomPosition(atom.GetIdx(), (pos.x, pos.y, pos.z))
    export_mol = rebuild_hydrogens_from_heavy_geometry(mol)
    issues = geometry_qc_issues(export_mol)
    if issues:
        raise RuntimeError("Export geometry quality check failed: " + "; ".join(issues))
    return export_mol


def rebuild_hydrogens_from_heavy_geometry(mol: Chem.Mol) -> Chem.Mol:
    heavy_mol = Chem.RemoveHs(Chem.Mol(mol), sanitize=True)
    export_mol = Chem.AddHs(heavy_mol, addCoords=True)
    assign_hydrogen_residue_info(export_mol)
    return export_mol


def assign_hydrogen_residue_info(mol: Chem.Mol) -> None:
    residue_counts: dict[tuple[str, int, str], int] = {}
    for atom in mol.GetAtoms():
        if atom.GetSymbol() != "H" or not atom.GetNeighbors():
            continue
        parent = atom.GetNeighbors()[0]
        parent_info = parent.GetPDBResidueInfo()
        if parent_info is None:
            continue
        key = (
            parent_info.GetChainId() or "A",
            parent_info.GetResidueNumber(),
            parent_info.GetResidueName().strip(),
        )
        residue_counts[key] = residue_counts.get(key, 0) + 1
        hydrogen_info = Chem.AtomPDBResidueInfo()
        hydrogen_info.SetName(f"H{residue_counts[key]:<3}"[:4])
        hydrogen_info.SetResidueName(parent_info.GetResidueName())
        hydrogen_info.SetResidueNumber(parent_info.GetResidueNumber())
        hydrogen_info.SetChainId(parent_info.GetChainId() or "A")
        hydrogen_info.SetIsHeteroAtom(parent_info.GetIsHeteroAtom())
        atom.SetMonomerInfo(hydrogen_info)


def geometry_qc_issues(mol: Chem.Mol, max_issues: int = MAX_QC_ISSUES) -> tuple[str, ...]:
    if mol.GetNumConformers() == 0:
        return ("missing conformer",)

    conformer = mol.GetConformer()
    issues: list[str] = []
    for bond in mol.GetBonds():
        atom_a = bond.GetBeginAtom()
        atom_b = bond.GetEndAtom()
        distance = conformer.GetAtomPosition(atom_a.GetIdx()).Distance(conformer.GetAtomPosition(atom_b.GetIdx()))
        min_allowed, max_allowed = _bond_distance_limits(atom_a, atom_b)
        if distance < min_allowed or distance > max_allowed:
            issues.append(
                f"{_qc_atom_label(atom_a)}-{_qc_atom_label(atom_b)} {distance:.2f} A outside {min_allowed:.2f}-{max_allowed:.2f} A"
            )
            if len(issues) >= max_issues:
                break
    return tuple(issues)


def _bond_distance_limits(atom_a: Chem.Atom, atom_b: Chem.Atom) -> tuple[float, float]:
    symbols = {atom_a.GetSymbol(), atom_b.GetSymbol()}
    if "H" in symbols:
        return 0.70, 1.35
    if symbols == {"S"}:
        return 1.80, 2.20
    if "S" in symbols:
        return 1.20, 2.00
    return 1.00, 1.90


def _qc_atom_label(atom: Chem.Atom) -> str:
    info = atom.GetPDBResidueInfo()
    if info is None:
        return f"{atom.GetSymbol()}{atom.GetIdx() + 1}"
    return f"{info.GetResidueName().strip()}{info.GetResidueNumber()}:{info.GetName().strip()}"


def build_structure(record: PeptideRecord, settings: BuildSettings) -> BuiltStructure:
    sequence = validate_sequence(record.sequence)
    seed_mol = build_seed_molecule(sequence)
    built = minimize_with_openmm(seed_mol, sequence, settings)
    return BuiltStructure(
        record=record,
        pdb_block=built.pdb_block,
        rdkit_mol=built.rdkit_mol,
        energy_kj_mol=built.energy_kj_mol,
        variants=built.variants,
        force_field=built.force_field,
    )


def export_built_structure(built: BuiltStructure, settings: BuildSettings) -> tuple[Path, ...]:
    settings.output_dir.mkdir(parents=True, exist_ok=True)
    base = safe_file_stem(f"{built.record.serial}_{built.record.sequence}")
    files = []
    for fmt in settings.formats:
        fmt = fmt.lower().strip(".")
        path = settings.output_dir / f"{base}.{fmt}"
        if fmt == "pdb":
            path.write_text(built.pdb_block, encoding="utf-8", newline="\n")
        elif fmt == "sdf":
            write_sdf(path, built.rdkit_mol, built)
        elif fmt == "mol2":
            path.write_text(mol_to_mol2(built.rdkit_mol, base), encoding="utf-8", newline="\n")
        elif fmt == "pdbqt":
            path.write_text(mol_to_pdbqt(built.rdkit_mol), encoding="utf-8", newline="\n")
        else:
            raise ValueError(f"Unsupported export format: {fmt}")
        files.append(path)
    return tuple(files)


def build_and_export(record: PeptideRecord, settings: BuildSettings) -> BuildResult:
    try:
        built = build_structure(record, settings)
        files = export_built_structure(built, settings)
        return BuildResult(record=record, ok=True, files=files, energy_kj_mol=built.energy_kj_mol, message="OK")
    except Exception as exc:
        return BuildResult(record=record, ok=False, message=str(exc))


def safe_file_stem(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]", "_", value)


def write_sdf(path: Path, mol: Chem.Mol, built: BuiltStructure) -> None:
    export_mol = Chem.Mol(mol)
    export_mol.SetProp("_Name", f"{built.record.serial}_{built.record.sequence}")
    export_mol.SetProp("Sequence", built.record.sequence)
    export_mol.SetProp("ForceField", built.force_field)
    export_mol.SetProp("OpenMM_Energy_kJ_mol", f"{built.energy_kj_mol:.6f}")
    writer = Chem.SDWriter(str(path))
    try:
        writer.write(export_mol)
    finally:
        writer.close()


def mol_to_pdbqt(mol: Chem.Mol) -> str:
    try:
        from meeko import MoleculePreparation, PDBQTWriterLegacy
    except ImportError as exc:
        raise RuntimeError("PDBQT export requires meeko and gemmi.") from exc

    prep = MoleculePreparation()
    setups = prep.prepare(mol)
    if not setups:
        raise RuntimeError("Meeko could not prepare this peptide for PDBQT export.")
    pdbqt_string, ok, error = PDBQTWriterLegacy.write_string(setups[0], bad_charge_ok=True)
    if not ok:
        raise RuntimeError(error or "Meeko PDBQT writer failed.")
    return pdbqt_string


def tripos_type(atom: Chem.Atom) -> str:
    symbol = atom.GetSymbol()
    hybridization = atom.GetHybridization()
    aromatic = atom.GetIsAromatic()
    if symbol == "C":
        if aromatic:
            return "C.ar"
        if hybridization == rdchem.HybridizationType.SP2:
            return "C.2"
        return "C.3"
    if symbol == "N":
        if aromatic:
            return "N.ar"
        if hybridization == rdchem.HybridizationType.SP2:
            return "N.am"
        return "N.3"
    if symbol == "O":
        if atom.GetFormalCharge() < 0:
            return "O.co2"
        if hybridization == rdchem.HybridizationType.SP2:
            return "O.2"
        return "O.3"
    if symbol == "S":
        return "S.3"
    if symbol == "H":
        return "H"
    return symbol


BOND_TYPES = {
    Chem.rdchem.BondType.SINGLE: "1",
    Chem.rdchem.BondType.DOUBLE: "2",
    Chem.rdchem.BondType.TRIPLE: "3",
    Chem.rdchem.BondType.AROMATIC: "ar",
}


def mol_to_mol2(mol: Chem.Mol, name: str) -> str:
    export_mol = Chem.Mol(mol)
    try:
        AllChem.ComputeGasteigerCharges(export_mol)
    except Exception:
        pass

    conformer = export_mol.GetConformer()
    lines = [
        "@<TRIPOS>MOLECULE",
        name,
        f"{export_mol.GetNumAtoms()} {export_mol.GetNumBonds()} 0 0 0",
        "SMALL",
        "GASTEIGER",
        "",
        "@<TRIPOS>ATOM",
    ]

    element_counts: dict[str, int] = {}
    for atom in export_mol.GetAtoms():
        info = atom.GetPDBResidueInfo()
        atom_name = info.GetName().strip() if info is not None else ""
        if not atom_name:
            element_counts[atom.GetSymbol()] = element_counts.get(atom.GetSymbol(), 0) + 1
            atom_name = f"{atom.GetSymbol()}{element_counts[atom.GetSymbol()]}"
        residue_number = info.GetResidueNumber() if info is not None else 1
        residue_name = info.GetResidueName().strip() if info is not None else name[:8]
        pos = conformer.GetAtomPosition(atom.GetIdx())
        charge = 0.0
        if atom.HasProp("_GasteigerCharge"):
            charge = atom.GetDoubleProp("_GasteigerCharge")
            if math.isnan(charge) or math.isinf(charge):
                charge = 0.0
        lines.append(
            f"{atom.GetIdx() + 1:7d} {atom_name:<8s} {pos.x:10.4f} {pos.y:10.4f} {pos.z:10.4f} "
            f"{tripos_type(atom):<8s} {residue_number:4d} {residue_name:<8s} {charge:10.4f}"
        )

    lines.append("@<TRIPOS>BOND")
    for index, bond in enumerate(export_mol.GetBonds(), start=1):
        lines.append(
            f"{index:6d} {bond.GetBeginAtomIdx() + 1:6d} {bond.GetEndAtomIdx() + 1:6d} "
            f"{BOND_TYPES.get(bond.GetBondType(), '1'):>4s}"
        )
    return "\n".join(lines) + "\n"


def estimate_sequence_properties(sequence: str) -> dict[str, float | int | str]:
    sequence = validate_sequence(sequence)
    mol = Chem.MolFromSequence(sequence)
    if mol is None:
        return {"sequence": sequence, "length": len(sequence)}
    exact_weight = Descriptors.ExactMolWt(mol)
    return {
        "sequence": sequence,
        "length": len(sequence),
        "exact_mw": round(float(exact_weight), 4),
        "aromatic_count": sum(1 for aa in sequence if aa in "FWYH"),
        "acidic_count": sum(1 for aa in sequence if aa in "DE"),
        "basic_count": sum(1 for aa in sequence if aa in "KRH"),
    }
